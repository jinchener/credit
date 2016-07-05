#!/usr/bin/python27
# -*- coding: utf-8 -*-
import xlrd
import sys
reload(sys)
sys.setdefaultencoding("utf-8")
from openpyxl import load_workbook



data = xlrd.open_workbook('act.xlsx')
table = data.sheets()[0]
#table.col_values(4)
nrows1 = table.nrows
print nrows1
word_list=table.col_values(0)
#for i in word_list:
#  print i



wb = load_workbook(filename=r'work.xlsx')

sheets = wb.get_sheet_names()   # 获取所有表格(worksheet)的名字
sheet0 = sheets[0]  # 第一个表格的名称
ws = wb.get_sheet_by_name(sheet0)
nrows = len(ws.rows)

'''
workdata=xlrd.open_workbook('work.xlsx')
table1=workdata.sheets()[0]
nrows=table1.nrows
#workdatalist=table1.col_values(23)
'''

word_dict = dict.fromkeys(word_list,0)

f = open("outfenci.txt", "w")
print nrows
for i in range(nrows):
    #cell_4 = table.cell(i, 23).value
    cell_title=ws.cell(row=i+1,column=18).value
    cell_content=ws.cell(row=i+1, column=24).value
    if cell_content or cell_title:
      for item in word_list:
          if (item in cell_title) or (item in cell_content):
            word_dict[item]+=1
          else:
            pass
      print 'completed schedule: %.2f'%((i+1)*100.00/nrows)+'%'
      sys.stdout.flush()

word_dictll=sorted(word_dict.items(), key=lambda d: d[1], reverse = True)
for i in word_dictll:
    f.write(i[0]+' '+str(i[1]))
    f.write('\n')

f.close()
raw_input()