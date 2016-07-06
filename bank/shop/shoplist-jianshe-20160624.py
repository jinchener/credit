#!/usr/bin/env python
# -*-coding: utf-8 -*-
import json
import openpyxl
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
import requests
import urllib2
from bs4 import BeautifulSoup as bs
import sys
import cookielib
import traceback
import chardet

reload(sys)
sys.setdefaultencoding("utf-8")


def get_request(url):
    # print "get: "+url+", refer:"+referUrl
    # 可以设置超时
    #socket.setdefaulttimeout(30)
    #sleepSec = random.randrange(6, 10)
    #time.sleep(sleepSec)
    # 可以加入参数	[无参数，使用get，以下这种方式，使用post]
    params = {"wd": "a", "b": "2"}
    enable_proxy = True
    proxy = urllib2.ProxyHandler({"http": "http://222.95.249.189:8090"})
    proxy_support = urllib2.ProxyHandler({})
    opener = urllib2.build_opener(proxy)
    urllib2.install_opener(opener)
    # 可以加入请求头信息，以便识别
    i_headers = {"User-Agent": "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1) Gecko/20090624 Firefox/3.5",
                 "Accept": "*/*"}
    # 设置保存cookie的文件，同级目录下的cookie.txt
    filename = 'cookie.txt'
    # 声明一个MozillaCookieJar对象实例来保存cookie，之后写入文件
    cookie = cookielib.MozillaCookieJar(filename)
    # 利用urllib2库的HTTPCookieProcessor对象来创建cookie处理器
    handler = urllib2.HTTPCookieProcessor(cookie)
    # 通过handler来构建opener
    opener = urllib2.build_opener(handler)
    # 创建一个请求，原理同urllib2的urlopen
    response = opener.open("http://sh.lianjia.com/xiaoqu/")
    # 保存cookie到文件
    cookie.save(ignore_discard=True, ignore_expires=True)
    # use post,have some params post to server,if not support ,will throw exception
    # req = urllib2.Request(url, data=urllib.urlencode(params), headers=i_headers)
    req = urllib2.Request(url, headers=i_headers)

    # 创建request后，还可以进行其他添加,若是key重复，后者生效
    # request.add_header('Accept','application/json')
    # 可以指定提交方式
    # request.get_method = lambda: 'PUT'
    try_count =0
    while 1:
        if try_count >= 10:
            print 'nothing'
            break
        try:
            opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cookie))
            page = opener.open(req)
            data = page.read()
            pageUrl = page.geturl()
            page.close()
            break
        except urllib2.HTTPError, e:
            print "Error Code:", e.code
            data = ''
            continue
        except urllib2.URLError, e:
            print "Error Reason:", e.reason
            data = ''
            continue
        except:
            print traceback.format_exc()
            data = ''
            continue
    return data


with open('city', 'r') as cityfile:
    # citys=json.loads(cityfile)
    cityjs = cityfile.readline().strip()
    # print type(cityjs)
    # print cityjs
    cityss = json.loads(cityjs, encoding="GB2312")["creditcard_city"]
    wb = Workbook()  # 创建工作薄
    ew = ExcelWriter(workbook=wb)  # 写入工作薄对象
    ws = wb.worksheets[0]  # 默认第一个表格
    ws.title = "data"
    list = ["prov_name","city_name","biz_id", "biz_name", "cate_id", "cate_name", "catechild_id", "catechild_name", "province_id", "province",
            "city_id", "city", "biz_addr", "biz_desc", "start_level", "is_new", "is_hot", "smallimage", "life_id",
            "life", "biz_phone", "biz_cmsg", "biz_image"]
    cellnum=1
    for i in range(len(list)):
        ws.cell(row=cellnum , column=i+1).value = list[i]
    ew.save(filename='result.xlsx')
    cellnum+=1

    for prov in cityss:
        prov_name= prov["prov_name"]
        prov_code= prov["prov_code"]
        citys=prov["citys"]
        for city in citys:
            city_name = city["city_name"]
            city_code = city["city_code"]
            branch_code = city["branch_code"]
            url="http://creditcard.ccb.com/webtran/get_crd_info.gsp?table_type=2&card_province=%s&card_city=%s&startNum=1&endNum=500"%(prov_code,city_code)
            data=get_request(url)
            try:
                datajs = json.loads(data, encoding="GB2312")
            except UnicodeDecodeError:
                print chardet.detect(data)
                #dataJsonStrUni = data.decode("GB2312");
                #datajs = json.loads(dataJsonStrUni, encoding="GB2312");
            except ValueError:
                print data


            totalNum=datajs["totalNum"]
            for i in range((int(totalNum)/500)+1):
                if not i:
                    obj=datajs["obj"]
                    for j in obj:
                        ws.cell(row=cellnum, column=1).value =prov_name
                        ws.cell(row=cellnum, column=2).value = city_name
                        for i in range(len(list[2:])):
                            ws.cell(row=cellnum, column=i+3).value = j[list[i+2]]
                        cellnum+=1
                        print cellnum
                    ew.save(filename='result.xlsx')
                else:
                    urltmp= "http://creditcard.ccb.com/webtran/get_crd_info.gsp?table_type=2&card_province=%s&card_city=%s&startNum=%s&endNum=%s" % (prov_code, city_code,(500*i+1),500*(i+1))
                    data = get_request(url)
                    datajs = json.loads(data, encoding="GB2312")
                    obj = datajs["obj"]
                    for j in obj:
                        ws.cell(row=cellnum, column=1).value = prov_name
                        ws.cell(row=cellnum, column=2).value = city_name
                        for i in range(len(list[2:])):
                            ws.cell(row=cellnum, column=i + 3).value = j[list[i + 2]]
                        cellnum += 1
                        print cellnum
                    ew.save(filename='result.xlsx')



    '''
    wb = Workbook()  # 创建工作薄
    ew = ExcelWriter(workbook=wb)  # 写入工作薄对象
    ws = wb.worksheets[0]  # 默认第一个表格
    ws.title = "data"
    for i in range(len(list)):
        ilist = list[i].strip().split('^')
        for j in range(len(ilist)):

        linenum = linenum + 1
        print linenum, c
    ew.save(filename=finename)
    '''
