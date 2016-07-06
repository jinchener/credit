#!/usr/bin/env python
# -*-coding: utf-8 -*-
import json
import openpyxl
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
import requests
import urllib2

from bs4 import BeautifulSoup as bs
import sys
import cookielib
import traceback
from urlparse import urljoin
from urlparse import urlparse
from urlparse import urlunparse
from posixpath import normpath
from lxml import etree
import chardet
import httplib
#httplib.HTTPConnection._http_vsn = 10
#httplib.HTTPConnection._http_vsn_str = 'HTTP/1.0'


reload(sys)
sys.setdefaultencoding("utf8")


def myjoin(base, url):
    url1 = urljoin(base, url)
    arr = urlparse(url1)
    path = normpath(arr[2])
    return urlunparse((arr.scheme, arr.netloc, path, arr.params, arr.query, arr.fragment))

def get_request(url):
    # print "get: "+url+", refer:"+referUrl
    # 可以设置超时
    #socket.setdefaulttimeout(30)
    #sleepSec = random.randrange(6, 10)
    #time.sleep(sleepSec)
    # 可以加入参数	[无参数，使用get，以下这种方式，使用post]
    params = {"wd": "a", "b": "2"}
    enable_proxy = True
    proxy = urllib2.ProxyHandler({"http": "http://222.95.249.189:8090"})
    proxy_support = urllib2.ProxyHandler({})
    opener = urllib2.build_opener(proxy)
    urllib2.install_opener(opener)
    # 可以加入请求头信息，以便识别
    i_headers = {"User-Agent": "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1) Gecko/20090624 Firefox/3.5",
                 "Accept": "*/*"}
    # 设置保存cookie的文件，同级目录下的cookie.txt
    filename = 'cookie.txt'
    # 声明一个MozillaCookieJar对象实例来保存cookie，之后写入文件
    cookie = cookielib.MozillaCookieJar(filename)
    # 利用urllib2库的HTTPCookieProcessor对象来创建cookie处理器
    handler = urllib2.HTTPCookieProcessor(cookie)
    # 通过handler来构建opener
    opener = urllib2.build_opener(handler)
    # 创建一个请求，原理同urllib2的urlopen
    response = opener.open("http://creditcard.cmbc.com.cn/promotioninfo/List.aspx?CityId=0&page=1")
    # 保存cookie到文件
    cookie.save(ignore_discard=True, ignore_expires=True)
    # use post,have some params post to server,if not support ,will throw exception
    # req = urllib2.Request(url, data=urllib.urlencode(params), headers=i_headers)
    req = urllib2.Request(url, headers=i_headers)

    # 创建request后，还可以进行其他添加,若是key重复，后者生效
    # request.add_header('Accept','application/json')
    # 可以指定提交方式
    # request.get_method = lambda: 'PUT'
    try_count = 0
    while 1:
        if try_count >= 10:
            print 'nothing'
        try:
            opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cookie))
            page = opener.open(req)
            data = page.read()
            pageUrl = page.geturl()
            page.close()
            break
        except urllib2.HTTPError, e:
            try_count += 1
            print "Error Code:", e.code
            data = ''
            continue
        except urllib2.URLError, e:
            try_count += 1
            print "Error Reason:", e.reason
            data = ''
            continue
        except:
            try_count += 1
            print traceback.format_exc()
            data = ''
            continue
    return data


nation='http://www.icbc.com.cn/icbc/%E7%89%A1%E4%B8%B9%E5%8D%A1/%E7%BC%A4%E7%BA%B7%E6%B4%BB%E5%8A%A8/%E5%9C%B0%E5%8C%BA%E4%BC%98%E6%83%A0%E6%B4%BB%E5%8A%A8%E5%88%97%E8%A1%A8.htm'
base='http://www.icbc.com.cn/icbc/%E7%89%A1%E4%B8%B9%E5%8D%A1/%E7%BC%A4%E7%BA%B7%E6%B4%BB%E5%8A%A8/%E5%9C%B0%E5%8C%BA%E4%BC%98%E6%83%A0%E6%B4%BB%E5%8A%A8%E5%88%97%E8%A1%A8-'
first='http://www.cmbc.com.cn/'
#datas=get_request(nation)
#chardit=chardet.detect(datas)
#print chardit
#pages = etree.HTML(datas.decode('gb2312'))#.decode('utf-8')
#print datas.decode('gb2312')
#pagenum=pages.xpath("/html/body/div[1]/form/div/div[3]/div/div[3]/table/tbody/tr/td/table/tbody/tr/td[1]/div")
#print pagenum,type(pagenum),len(pagenum)
pagenum=5
wb = Workbook()  # 创建工作薄
ew = ExcelWriter(workbook=wb)  # 写入工作薄对象
ws = wb.worksheets[0]  # 默认第一个表格
ws.title = "data"
cellnum=1
list=[u"名称",u"发布日期",u"url"]
for i in range(len(list)):
    ws.cell(row=cellnum, column=i+1).value = list[i]
cellnum += 1


for i in range(int(pagenum)):
    if i == 0:
        url=nation
        try:
            data = str(get_request(url).decode('gb2312'))
            # print data
            page = etree.HTML(data.decode('utf-8'))  # .decode('utf-8')
        except UnicodeDecodeError:
            data = get_request(url)
            chardit = chardet.detect(data)
            page = etree.HTML(data.decode('utf-8'))
            print chardit

        name = page.xpath("//*[@id=\"detailCon\"]/div/div[2]/div/table//tr/td[2]/span/a")
        #print  name,type(name),len(name)
        #area=page.xpath("/html/body/div[1]/form/div/div[3]/div/div[2]/table//tr/td/ul/li/em[1]")
        #date=page.xpath("//*[@id=\"Main11\"]/div/div[1]/div/div/table//tr/td[3]/span")
        for j in range(len(name)):
            act=name[j].text.strip()
            #time=date[j].text.strip()
            time=''
            #city=area[j].text.strip()
            llend=name[j].get("href").strip()
            #print act,time
            href=myjoin(nation,llend)
            ws.cell(row=cellnum, column=1).value = act
            ws.cell(row=cellnum, column=2).value = time
            #ws.cell(row=cellnum, column=3).value = city
            ws.cell(row=cellnum, column=3).value = href
            cellnum += 1
            print cellnum
        ew.save(filename='result.xlsx')






    else:
        urlend="PageList-%s.htm"%(pagenum-i)
        url=base+urlend
        try:
            data = str(get_request(url).decode('gb2312'))
            # print data
            page = etree.HTML(data.decode('utf-8'))  # .decode('utf-8')
        except UnicodeDecodeError:
            data=get_request(url)
            chardit = chardet.detect(data)
            page = etree.HTML(data.decode('utf-8'))
            print chardit

        name = page.xpath("//*[@id=\"detailCon\"]/div/div[2]/div/table//tr/td[2]/span/a")
        # print  name,type(name),len(name)
        # area=page.xpath("/html/body/div[1]/form/div/div[3]/div/div[2]/table//tr/td/ul/li/em[1]")
        # date=page.xpath("//*[@id=\"Main11\"]/div/div[1]/div/div/table//tr/td[3]/span")
        for j in range(len(name)):
            act = name[j].text.strip()
            #time = date[j].text.strip()
            time=''
            #city = area[j].text.strip()
            llend = name[j].get("href").strip()
            href = myjoin(nation, llend)
            ws.cell(row=cellnum, column=1).value = act
            ws.cell(row=cellnum, column=2).value = time
            #ws.cell(row=cellnum, column=3).value = city
            ws.cell(row=cellnum, column=3).value = href
            cellnum += 1
            print cellnum
        ew.save(filename='result.xlsx')



#print nation+urlend





'''
with open('work', 'r') as cityfile:
    # citys=json.loads(cityfile)
    cityjs = cityfile.readline().strip()
    # print type(cityjs)
    # print cityjs
    citys = json.loads(cityjs, encoding="GB2312")
    # print type(cityjs)
    objs = citys["list"]
    wb = Workbook()  # 创建工作薄
    ew = ExcelWriter(workbook=wb)  # 写入工作薄对象
    ws = wb.worksheets[0]  # 默认第一个表格
    ws.title = "data"
    for cellnum, obj in enumerate(objs):
        title = obj["title"]
        # print active_id,type(active_id)
        titleLink = obj["titleLink"]
        end_date = obj["end_date"]
        active_name = obj["active_name"]
        url = 'http://ebank.ccb.com' + active_id + '.html'
        respose = get_request(url)
        if respose:
            soup = bs(respose, "lxml")
            content = soup.find_all("div", "content")[1]
            cont = content.find_all("p")
            person = cont[0].get_text()[6:]
            time = cont[1].get_text()[6:]
            detail = cont[3].get_text()
        else:
            person=''
            time=''
            detail=''

        #rl1=cont[3].find('a').get('href')
        #print url1

        ws.cell(row=cellnum + 1, column=1).value = active_id
        ws.cell(row=cellnum + 1, column=2).value = begin_date
        ws.cell(row=cellnum + 1, column=3).value = end_date
        ws.cell(row=cellnum + 1, column=4).value = active_name
        ws.cell(row=cellnum + 1, column=5).value = url
        ws.cell(row=cellnum + 1, column=6).value = person
        ws.cell(row=cellnum + 1, column=7).value = time
        ws.cell(row=cellnum + 1, column=8).value = detail
        #ws.cell(row=cellnum + 1, column=9).value = url1




        print cellnum

    ew.save(filename='result.xlsx')
    # print citys
    # print type(citys)

'''
'''
    wb = Workbook()  # 创建工作薄
    ew = ExcelWriter(workbook=wb)  # 写入工作薄对象
    ws = wb.worksheets[0]  # 默认第一个表格
    ws.title = "data"
    for i in range(len(list)):
        ilist = list[i].strip().split('^')
        for j in range(len(ilist)):

        linenum = linenum + 1
        print linenum, c
    ew.save(filename=finename)
'''
