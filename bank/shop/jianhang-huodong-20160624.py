#!/usr/bin/env python
# -*-coding: utf-8 -*-
import json
import openpyxl
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
import requests
import urllib2
from bs4 import BeautifulSoup as bs
import sys
import cookielib
import traceback

reload(sys)
sys.setdefaultencoding("utf-8")


def get_request(url):
    # print "get: "+url+", refer:"+referUrl
    # 可以设置超时
    #socket.setdefaulttimeout(30)
    #sleepSec = random.randrange(6, 10)
    #time.sleep(sleepSec)
    # 可以加入参数	[无参数，使用get，以下这种方式，使用post]
    params = {"wd": "a", "b": "2"}
    enable_proxy = True
    proxy = urllib2.ProxyHandler({"http": "http://222.95.249.189:8090"})
    proxy_support = urllib2.ProxyHandler({})
    opener = urllib2.build_opener(proxy)
    urllib2.install_opener(opener)
    # 可以加入请求头信息，以便识别
    i_headers = {"User-Agent": "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9.1) Gecko/20090624 Firefox/3.5",
                 "Accept": "*/*"}
    cookie = cookielib.MozillaCookieJar()
    cookie.load('cookie.txt', ignore_discard=True, ignore_expires=True)
    # use post,have some params post to server,if not support ,will throw exception
    # req = urllib2.Request(url, data=urllib.urlencode(params), headers=i_headers)
    req = urllib2.Request(url, headers=i_headers)

    # 创建request后，还可以进行其他添加,若是key重复，后者生效
    # request.add_header('Accept','application/json')
    # 可以指定提交方式
    # request.get_method = lambda: 'PUT'
    try:
        opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cookie))
        page = opener.open(req)
        data = page.read()
        pageUrl = page.geturl()
        page.close()
        return data
    except urllib2.HTTPError, e:
        print "Error Code:", e.code
        data=''
        return data
    except urllib2.URLError, e:
        print "Error Reason:", e.reason
    except:
        print traceback.format_exc()


with open('work', 'r') as cityfile:
    # citys=json.loads(cityfile)
    cityjs = cityfile.readline().strip()
    # print type(cityjs)
    # print cityjs
    citys = json.loads(cityjs, encoding="GB2312")
    # print type(cityjs)
    objs = citys["obj"]
    wb = Workbook()  # 创建工作薄
    ew = ExcelWriter(workbook=wb)  # 写入工作薄对象
    ws = wb.worksheets[0]  # 默认第一个表格
    ws.title = "data"
    for cellnum, obj in enumerate(objs):
        active_id = obj["active_id"]
        # print active_id,type(active_id)
        begin_date = obj["begin_date"]
        end_date = obj["end_date"]
        active_name = obj["active_name"]
        url = 'http://creditcard.ccb.com/cn/creditcard/acitivity/' + active_id + '.html'
        respose = get_request(url)
        if respose:
            soup = bs(respose, "lxml")
            content = soup.find_all("div", "content")[1]
            cont = content.find_all("p")
            person = cont[0].get_text()[6:]
            time = cont[1].get_text()[6:]
            detail = cont[3].get_text()
        else:
            person=''
            time=''
            detail=''

        #rl1=cont[3].find('a').get('href')
        #print url1

        ws.cell(row=cellnum + 1, column=1).value = active_id
        ws.cell(row=cellnum + 1, column=2).value = begin_date
        ws.cell(row=cellnum + 1, column=3).value = end_date
        ws.cell(row=cellnum + 1, column=4).value = active_name
        ws.cell(row=cellnum + 1, column=5).value = url
        ws.cell(row=cellnum + 1, column=6).value = person
        ws.cell(row=cellnum + 1, column=7).value = time
        ws.cell(row=cellnum + 1, column=8).value = detail
        #ws.cell(row=cellnum + 1, column=9).value = url1




        print cellnum

    ew.save(filename='result.xlsx')
    # print citys
    # print type(citys)

    '''
    wb = Workbook()  # 创建工作薄
    ew = ExcelWriter(workbook=wb)  # 写入工作薄对象
    ws = wb.worksheets[0]  # 默认第一个表格
    ws.title = "data"
    for i in range(len(list)):
        ilist = list[i].strip().split('^')
        for j in range(len(ilist)):

        linenum = linenum + 1
        print linenum, c
    ew.save(filename=finename)
    '''
